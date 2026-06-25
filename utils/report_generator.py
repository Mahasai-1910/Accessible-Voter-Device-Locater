"""Report generation for Excel, CSV, and PDF exports."""

from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from utils.data_loader import add_compliance_scores
from utils.theme import LOGO_PATH, ROOT

REPORTS_DIR = ROOT / "reports"


def _ensure_reports_dir() -> Path:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    return REPORTS_DIR


def generate_compliance_report(locations: pd.DataFrame) -> pd.DataFrame:
    df = add_compliance_scores(locations)
    cols = [
        "Location ID",
        "Location Name",
        "City",
        "State",
        "Compliance Score",
        *[
            c
            for c in [
                "Wheelchair Access",
                "Braille Device",
                "Audio Ballot",
                "Large Print Ballot",
                "Sign Language Assistance",
                "Staff Assistance",
            ]
            if c in df.columns
        ],
    ]
    return df[[c for c in cols if c in df.columns]].sort_values("Compliance Score", ascending=False)


def generate_device_report(devices: pd.DataFrame) -> pd.DataFrame:
    if devices.empty:
        return devices
    return devices.sort_values(["Polling Location", "Device Type"])


def generate_location_summary(locations: pd.DataFrame) -> pd.DataFrame:
    df = add_compliance_scores(locations)
    if df.empty:
        return df
    summary = (
        df.groupby("State")
        .agg(
            Locations=("Location ID", "count"),
            Avg_Compliance=("Compliance Score", "mean"),
            Fully_Accessible=("Compliance Score", lambda x: (x >= 85).sum()),
        )
        .reset_index()
        .round(1)
    )
    return summary


def generate_assistance_report(requests: pd.DataFrame) -> pd.DataFrame:
    if requests.empty:
        return requests
    return requests.sort_values("Date", ascending=False)


def export_excel_report(
    report_df: pd.DataFrame,
    sheet_name: str = "Report",
    title: str = "CAP AI Report",
) -> bytes:
    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_fill = PatternFill("solid", fgColor="0047AB")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(color="0047AB", bold=True, size=16)

    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A3"] = "CAP AI — Accessible Voting Device Locator"

    start_row = 5
    if LOGO_PATH.exists():
        try:
            img = XLImage(str(LOGO_PATH))
            img.width = 80
            img.height = 80
            ws.add_image(img, "E1")
        except Exception:
            pass

    if not report_df.empty:
        headers = list(report_df.columns)
        for ci, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=ci, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for ri, row in enumerate(report_df.itertuples(index=False), start_row + 1):
            for ci, value in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=value)

        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_pdf_report(
    report_df: pd.DataFrame,
    title: str = "CAP AI Accessibility Report",
    subtitle: str = "Accessible Voting Device Locator",
) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.75 * inch, bottomMargin=0.75 * inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], textColor=colors.HexColor("#0047AB"), fontSize=18)
    sub_style = ParagraphStyle("Sub", parent=styles["Normal"], textColor=colors.HexColor("#00AEEF"), fontSize=11)
    elements = []

    if LOGO_PATH.exists():
        try:
            elements.append(Image(str(LOGO_PATH), width=1.2 * inch, height=1.2 * inch))
        except Exception:
            pass

    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(subtitle, sub_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 0.3 * inch))

    if not report_df.empty:
        data = [list(report_df.columns)] + report_df.astype(str).values.tolist()
        col_count = len(report_df.columns)
        col_width = min(7.0 / col_count, 1.5) * inch
        table = Table(data, colWidths=[col_width] * col_count, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0047AB")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ]
            )
        )
        elements.append(table)
    else:
        elements.append(Paragraph("No data available for this report.", styles["Normal"]))

    elements.append(Spacer(1, 0.5 * inch))
    elements.append(
        Paragraph(
            "© CAP AI — Accessible Voting Device Locator | WCAG 2.1 AA Compliant Platform",
            ParagraphStyle("Footer", parent=styles["Normal"], fontSize=8, textColor=colors.grey),
        )
    )

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def save_report_file(content: bytes, filename: str) -> Path:
    out_dir = _ensure_reports_dir()
    path = out_dir / filename
    path.write_bytes(content)
    return path
